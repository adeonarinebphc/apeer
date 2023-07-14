import seaborn as sns
import pandas as pd
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
import numpy as np
from sklearn.cluster import KMeans
from descartes import PolygonPatch
from json import dumps
import json
from collections import defaultdict
import csv
import openpyxl
import os

import matplotlib
matplotlib.use("Agg")

def FindMissingCols():

	# for some reason, counties only has 173 columns compared to the full list of 177
	# from tract data. missing Columns are:
	# ['1_BROMOPROPANE', '1_1_DIMETHYLHYDRAZINE', '1_1_2_TRICHLOROETHANE', '1_1_2_2_TETRACHLOROETHANE']
	cfile = "2018_Toxics_Ambient_Concentrations.county.tsv"
	chemdata_county = pd.read_csv(cfile, header=0, index_col=0, sep="\t")
	print(chemdata_county.columns)

	cfile = "2018_Toxics_Ambient_Concentrations.tract.tsv"
	chemdata_tract = pd.read_csv(cfile, header=0, index_col=0, sep="\t")
	print(chemdata_tract.columns)

	missing_cols = []
	for x in range(0, len(chemdata_tract.columns)):
		tcol = chemdata_tract.columns[x]
		if tcol not in chemdata_county.columns:
			print("Missing column: " + tcol)
	missing_cols.append(tcol)

	return missing_cols

def LoadMapJSON(tmode):

	# Map the data
	tfile = "us_counties.geojson"
	if tmode == "tract":
		tfile = "tracts.geojson"

	print("LoadMapJSON: " + tfile)
	with open(tfile, "r") as json_file:
		json_data = json.load(json_file)

	return json_data

def GetPopData():

	# get file
	tcnt = 0
	popdata = defaultdict(float)
	tfile = "EJSCREEN_2021_USPR_Tracts.csv"
	with open(tfile, "r") as infile:
		for line in infile:
			ldata = line.split(',')
			#print(str(line))
			if tcnt > 0:
				tfips = ldata[1]
				tpop = ldata[2]
				popdata[tfips] = float(tpop)
			tcnt += 1

	return popdata

def GetCountyData(popdata, tfile):
	
	# calculate county-level populations
	countypop = defaultdict(float)
	for tfips in popdata:
		nfips = tfips
		if len(nfips) == 10:
			nfips = "0" + nfips
		tcounty = nfips[0:5]
		countypop[tcounty] += popdata[tfips]
	
	# now convert the chemical data to county level
	lcnt = 0
	countydata = defaultdict(lambda: defaultdict(float))
	with open(tfile, "r") as infile:
		for line in infile:
			line = line.strip()
			ldata = line.split("\t")
			if lcnt == 0:
				theader = ldata
				#print("County header: " + str(theader))
				print("Debug 1b - Number of cols: " + str(len(theader)))
			if lcnt > 0:
				#tdesc = ldata[2]
				#if tdesc.find("Entire") == -1:
				tfips = ldata[0]
				if len(tfips) == 10:
					tfips = "0" + tfips
				tcounty = tfips[0:5]
				if tcounty in countypop:
					tweight = popdata[tfips] / countypop[tcounty]
					for x in range(1, len(theader)):
						nval = float(ldata[x]) * float(tweight)
						countydata[tcounty][theader[x]] += nval
			lcnt += 1
	infile.close()
	
	# write out county data
	print("Debug 1x - Number of columns: " + str(len(theader)))
	ofile = "2018_Toxics_Ambient_Concentrations.county.tsv"
	f = open(ofile, "w")
	tline = "FIPS\t"
	for x in range(1, len(theader)):
		print(str(x) + "\t" + theader[x])
		tline = tline + theader[x] + "\t"
	tline = tline.strip()
	f.write(tline + "\n")
	for tfips in countydata:
		tline = tfips + "\t"
		for x in range(1, len(theader)):
			tline = tline + str(countydata[tfips][theader[x]]) + "\t"
		tline = tline.strip()
		f.write(tline + "\n")
	f.close()
	
	return ofile
	

def RunPCA(pcadata, tfile, ttitle):

	pca = PCA()
	pcadata = pcadata.replace(np.nan, 0)
	
	Xt = pca.fit_transform(pcadata)
	#loadings = pca.components_
	#loadings = pca.components_.T * np.sqrt(pca.explained_variance_)
	
	# print PC variance explained
	exp_var_pca = pca.explained_variance_ratio_
	plt.close()
	#plt.bar(range(0,len(exp_var_pca)), exp_var_pca, alpha=0.5, align='center', label='Individual explained variance')
	#plt.savefig(tfile + ".variance.pdf")
	#plt.close()
	
	# print loadings
	#loadings = pca.components_.T * np.sqrt(pca.explained_variance_)
	#loading_matrix = pd.DataFrame(loadings, columns=['PC1', 'PC2'], index=pcadata.columns)
	#print(loading_matrix)

	#plt.rcParams["axes.grid"] = False	
	#plt.rcParams["figure.figsize"] = (8,8)

	#fig, ax = plt.subplots()
	#ax.patch.set_edgecolor('black')  
	#ax.patch.set_linewidth('1')
	#ax.set_facecolor("#FFF")
	#ax.scatter(Xt[:,0], Xt[:,1])
	#ax.set_xlabel("Principal Component 1")
	#ax.set_ylabel("Principal Component 2")
	#ax.set_title(ttitle)
	#plt.show()
	dropdata = []
	
	#plt.savefig(tfile)
	#plt.show()
	plt.close()

	# count number of clusters
	#wcss = []
	#for i in range(1, 21):
	#	kmeans_pca = KMeans(n_clusters = i)
	#	kmeans_pca.fit(Xt)
	#	wcss.append(kmeans_pca.inertia_)
		
	#plt.figure(figsize=(10,8))
	#plt.plot(range(1,21), wcss, marker='o', linestyle='--')
	#plt.title(ttitle + " WCSS Number of Clusters")
	#plt.xlabel('Number of Clusters')
	#plt.ylabel('Kmeans with PCA Clustering')
	#plt.savefig(tfile + ".numclust.pdf")
	#plt.show()
	#plt.close()

	return Xt

def PCA_Kmeans(Xt_pca, numclust, pcadata, pcafile, clustfile, clust_tsv):

	cluster_colors = ["red", "green", "blue", "purple", "yellow", "violet", "orange"]

	PCA_components = pd.DataFrame(Xt_pca)
	model = KMeans(n_clusters = numclust)
	model.fit(PCA_components.iloc[:,:2])
	labels = model.predict(PCA_components.iloc[:,:2])
	
	colorlist = []
	for x in range(0, len(labels)):
		colorlist.append(cluster_colors[labels[x]])
	
	plt.close()
	#plt.scatter(PCA_components[0], PCA_components[1], c=colorlist)
	#plt.xlabel("PC1")
	#plt.ylabel("PC2")
	#plt.title("SDI -Means Clustered PCA (k=" + str(numclust) + ")")
	covidtable = []
	covidindex = []
	clustertable = []
	for i, name in enumerate(pcadata.index):
	
		# combine with COVID-19 data
		tfips = str(name)
		if len(tfips) == 4:
			tfips = '0' + tfips
		if len(tfips) == 10:
			tfips = '0' + tfips		

		# list of clusters
		trow = []
		tclust = []
		tclust.append(tfips)
		tclust.append(labels[i])			
		clustertable.append(tclust)
			
	#plt.savefig(pcafile)
	#plt.close()
	
	clusterframe = pd.DataFrame(clustertable, columns=["county", "cluster"])
	clusterframe = clusterframe.set_index("county")
	clusterframe.to_csv(clust_tsv, sep="\t")

	# now do a violin plot of COVID-19 rates for each k-means cluster
	#https://stackoverflow.com/questions/64785394/seaborn-violin-plot-for-single-column-splitting-by-a-categorical-column

	#fig.show()
	#fig.savefig(clustfile)
	
	return clusterframe

def ShowBinaryMap(beltlist, mapfile, json_data, backcolor, forecolor):

	# set colors
	tcolors = ["#ff0000", "#00ff00", "#0000ff", "#6a0dad", "#ffff00"]

	plt.close()
	fig = plt.figure(figsize=(25, 16))
	ax = fig.gca()
	for titem in json_data["features"]:
		tfips = titem["properties"]["GEOID"]
		if int(tfips[0:1]) < 6:
			get_color = backcolor
			if tfips in beltlist:
				get_color = forecolor
			#get_color = tcolors[color_index]
			poly = titem["geometry"]
			if str(poly) != "None":
				#print(tfips + "\t" + str(get_color))
				ax.add_patch(PolygonPatch(poly, fc=get_color, ec=get_color, alpha=1, zorder=2))
	
	#ax.axis('scaled')
	#ax.set_xlim(-180, -60)
	#ax.set_ylim(20, 80)
	ax.set_xlim(-130, -60)
	ax.set_ylim(23, 50)
	ax.set_facecolor('xkcd:white')
	#plt.xlabel("Longitude")
	#plt.ylabel("Latitude")

	fig.tight_layout()
	plt.axis("off")
	plt.savefig(mapfile)

def ShowClustersOnMap(tmode, datatable, numclust, tfile, json_data):

	# make filenames
	#bpath = "data"
	pcafile = tfile + ".pca.pdf"
	kfile = tfile + ".kmeans.pca.pdf"
	cluster_file = tfile + ".kmeans.pca.clusters.pdf"
	cluster_tsv = tfile + "." + str(numclust) + ".cluster.tsv"
	mapfile = tfile + ".map.png"
	
	# PCA
	datatable = datatable.replace(np.nan, 0)
	Xt_epa = RunPCA(datatable, pcafile, tfile)
	epa_clust = PCA_Kmeans(Xt_epa, numclust, datatable, kfile, cluster_file, cluster_tsv)
	
	#print("Clusters:")
	#print(epa_clust)
	epa_clust.to_csv(cluster_tsv, sep="\t")

	# set colors
	#tcolors = ["#ff0000", "#00ff00", "#0000ff", "#6a0dad", "#ffff00"]	
	#tcolors = ["red", "green", "blue", "purple", "yellow", "violet", "orange"]

	plt.close()
	plt.axis("off")
	fig = plt.figure(figsize=(5.5, 3.5))
	ax = fig.gca()
	
	# get which cluster has more FIPS codes in it (larger in size) - set this as the gray background
	# and set the smaller cluster as blue
	light_gray = "#f5f5f5"
	blue_color = "blue"
	clust0 = 0
	clust1 = 0
	if numclust == 2:
		for tfips in epa_clust.index:
			tclust = epa_clust.loc[tfips, "cluster"]
			if (tclust == 0):
				clust0 += 1
			if (tclust == 1):
				clust1 += 1
		if (clust0 >= clust1):
			tcolors = ["#f5f5f5", "blue"]
		if (clust0 < clust1):
			tcolors = ["blue", "#f5f5f5"]
	
	# assign colors - stroke belt = red, other = gray
	# if iowa (fips=19) is in the cluster, assign it gray
	#tcolors = ["blue", "red", "green", "purple", "orange", "yellow"]
	if numclust == 3:
		tcolors = ["red", "blue", "green"]
	if numclust == 4:
		tcolors = ["blue", "red", "green", "purple", "orange", "#f5f5f5"]
		#tcolors = ["#f5f5f5", "#f5f5f5", "#f5f5f5", "blue"]
	if numclust == 5:
		tcolors = ["blue", "red", "green", "purple", "orange", "#f5f5f5"]
	#tcolors = ["red", "green"]
	for titem in json_data["features"]:
		tfips = titem["properties"]["GEOID"]
		if tfips in epa_clust.index:
			if (tfips[0:2] == "19") and ((len(tfips) == 5) or (len(tfips) == 11)):
				if epa_clust.loc[tfips, "cluster"] == 0:
					#tcolors = ["green", "red"]
					if numclust == 3:
						tcolors = ["red", "blue", "green"]
					if numclust == 4:
						tcolors = ["red", "green", "purple", "blue"]
					if numclust == 5:
						tcolors = ["red", "blue", "green", "purple", "orange", "yellow"]
	
	# now plot on map
	for titem in json_data["features"]:
		tfips = titem["properties"]["GEOID"]
		if tfips in epa_clust.index:			
			color_index = epa_clust.loc[tfips, "cluster"]
			get_color = tcolors[color_index]
			poly = titem["geometry"]
			if poly is not None:
				ax.add_patch(PolygonPatch(poly, fc=get_color, ec=get_color, alpha=1, zorder=2))
	
	#ax.set_xlim(-180, -60)
	#ax.set_ylim(20, 80)
	ax.set_xlim(-130, -60)
	ax.set_ylim(23, 50)
	ax.set_facecolor('xkcd:white')
	#plt.xlabel("Longitude")
	#plt.ylabel("Latitude")
	fig.tight_layout()
	plt.axis("off")

	fig.tight_layout()
	plt.savefig(mapfile, dpi=1200)

def LoadPollutionData():

	tcnt = 0
	ttable = []
	with open("2018_Toxics_Ambient_Concentrations.txt", encoding="latin-1") as infile:
		for line in infile:
			
			line = line.strip()
			ldata = line.split("\t")
			
			if tcnt == 0:
				theader = []
				for x in range(5, len(ldata)):
					ldata[x] = ldata[x].replace("-", "_")
					ldata[x] = ldata[x].replace(",", "_")
					ldata[x] = ldata[x].replace(" ", "_")
					ldata[x] = ldata[x].replace("'", "_")
					ldata[x] = ldata[x].replace("\"", "")
					ldata[x] = ldata[x].replace("(", "_")
					ldata[x] = ldata[x].replace(")", "_")
					ldata[x] = ldata[x].replace("[", "_")
					ldata[x] = ldata[x].replace("]", "_")
					ldata[x] = ldata[x].replace("__", "_")
					
					if ldata[x][-1] == "_":
						ldata[x] = ldata[x][:-1]
					
					theader.append(ldata[x])
				
			if tcnt > 0:
				trow = []
				tdesc = ldata[2]
				if tdesc.find("Entire") == -1:
					#print("Description: " + tdesc)
					trow = [ldata[4]]
					for x in range(5, len(ldata)):
						trow.append(ldata[x])
					ttable.append(trow)
					#print(str(trow))
					
			tcnt += 1
				
	infile.close()

	#workbook = openpyxl.load_workbook("2018_Toxics_Ambient_Concentrations.xlsx")
	#worksheet = workbook.active
	
	#ttable = []
	#theader = []
	#for trow in range(1, worksheet.max_row + 1):
	
	#	rowdata = []
	#	if trow == 1:
	#		for tcol in range(6, worksheet.max_column + 1):
	#			tval = worksheet.cell(1, tcol).value
	#			#print("Header column: " + tval)
	#			theader.append(tval)
	#		
	#	if trow > 1:
	#		tdesc = worksheet.cell(trow, 3).value
	#		tfips = worksheet.cell(trow, 5).value
	#		rowdata.append(tfips)
	#		if tdesc.find("Entire") == -1:
	#			for tcol in range(6, worksheet.max_column + 1):
	#				nvalue = worksheet.cell(trow, tcol).value
	#				rowdata.append(nvalue)
	#			ttable.append(rowdata)
	#	
	#	if ((trow % 1000) == 0):
	#		print("Number of rows: " + str(trow))

	clist = ["fips"] + theader
	chemdata = pd.DataFrame(ttable, columns=clist)
	chemdata.set_index("fips", inplace=True)
	chemdata[theader] = chemdata[theader].apply(pd.to_numeric)
	chemdata = chemdata.replace(np.nan, 0)
	
	print(chemdata)

	return chemdata
			
def ShowCorrelations(corrmat, tfile):

	corrmat = chemdata.corr()
	sns.set(font_scale=0.2)
	cg = sns.clustermap(corrmat, cmap ="YlGnBu", linewidths = 0.1)
	cg.savefig(tfile, dpi=1200)

def PlotPairwiseMultiMaps(datalist, tmode, basedata, tbase, json_data):

	tcnt = 0
	filelist = []
	missing_chem = ['1_BROMOPROPANE', '1_1_DIMETHYLHYDRAZINE', '1_1_2_TRICHLOROETHANE', '1_1_2_2_TETRACHLOROETHANE']
	for x in range(0, len(datalist)):
		for y in range(x, len(datalist)):
			if (datalist[x] != datalist[y]):
				#print(str(tcnt) + "\t" + "Pair: " + datalist[x] + "," + datalist[y])
				tpair = [datalist[x], datalist[y]]
				tdata = basedata[tpair]
				ttag = datalist[x] + "-" + datalist[y]
				flist = []
				for tclust in range(2, 5):
					tfile = tbase + "." + ttag + "." + str(tclust) + ".pdf"
					#ShowClustersOnMap(tmode, tdata, tclust, tfile, json_data)
					# make filenames
					bpath = "data"
					pcafile = bpath + "//" + tfile + ".pca.pdf"
					kfile = bpath + "//" + tfile + ".kmeans.pca.pdf"
					cluster_file = bpath + "//" + tfile + ".kmeans.pca.clusters.pdf"
					cluster_tsv = bpath + "//" + tfile + "." + str(tclust) + ".cluster.tsv"
					mapfile = bpath + "//" + tfile + ".map.jpg"
				
					# PCA
					#tdata = tdata.replace(np.nan, 0)
					#Xt_epa = RunPCA(tdata, pcafile, tfile)
					#epa_clust = PCA_Kmeans(Xt_epa, tclust, tdata, kfile, cluster_file)
					#epa_clust.to_csv(cluster_tsv, sep="\t")
					#if not os.path.isfile(cluster_tsv):
					if datalist[x] in missing_chem:
						print("File - " + cluster_tsv)
						ShowClustersOnMap("county", tdata, tclust, tfile, json_data)

					flist.append(tfile + ".map.jpg")

				tcnt += 1
				
				filelist.append(flist)

def PlotMultiMaps(datalist, tmode, basedata, tbase, json_data):

	tcnt = 0
	filelist = []
	for x in range(0, len(datalist)):
		for y in range(x+1, len(datalist)):
			print("Pair: " + datalist[x] + "," + datalist[y])
			tpair = [datalist[x], datalist[y]]
			tdata = basedata[tpair]
			ttag = datalist[x] + "-" + datalist[y]
			flist = []
			for tclust in range(2, 6):
				tfile = tbase + "." + ttag + "." + str(tclust) + ".pdf"
				ShowClustersOnMap(tmode, tdata, tclust, tfile, json_data)
				flist.append(tfile + ".map.jpg")
			filelist.append(flist)
	
	# transpose table using zip
	ttable = [list(i) for i in zip(*filelist)]

	# get header
	headerlist = []
	for tcol in range(0, len(ttable[0])):
		tdata = ttable[0][tcol].split('.')
		tdata2 = tdata[1].split('-')
		tval1 = tdata2[0]
		tval2 = tdata2[1]
		nheader = tval1 + ',<br>' + tval2
		headerlist.append(nheader)

	tHTML = "<html>\n<body style=\"font-family: arial; font-size: 10px;\">\n<table style=\"width: 100%;\">\n"
	f = open(tbase + ".html", "w")
	tline = "\t<tr>\n\t\t<td></td>\n"
	for theader in headerlist:
		nheader = theader.replace('_', ' ')
		tline += "\t\t<td><font style=\"font-size: 9px;\">" + nheader + "</font></td>\n"
	tline += "\t<tr>\n"
	tHTML += tline
	
	clustnum = 2
	for trow in range(0, len(ttable)):
		tline = "\t<tr>\n"
		tline += "\t\t<td><font style=\"font-size: 14px\">" + str(clustnum) + "</font></td>"
		for tcol in range(0, len(ttable[trow])):
			tline += "\t\t<td><img src=\"" + ttable[trow][tcol] + "\" style=\"width: 100px;\"></td>\n"
		tline += "\t</tr>\n"
		clustnum += 1
		tHTML += tline
	tHTML += "</table>\n</body>\n</html>"
	f.write(tHTML)
	f.close()

#########################
### Run main analysis ###
#########################

# STEP 1: you will need to download the following libraries:
#seaborn
#pandas
#sklearn
#matplotlib
#numpy
#descartes
#openpyxl

# STEP 2: download AirToxScreen data (2018) from the EPA website, and save it as a tsv 
# "2018_Toxics_Ambient_Concentrations.tract.tsv"
# Download the file from here: https://drive.google.com/file/d/1IAwhEqD-DuBShfXH3Q1uUcY2y2bckTr2/view?usp=share_link

# STEP 3: download EJSCREEN data (https://www.epa.gov/ejscreen/download-ejscreen-data) 
# and calculate the population at the tract level, saving the file as 
# "EJSCREEN_2021_USPR_Tracts.csv" (2nd column = FIPS code, 3rd column = population)
# Download the file from here: https://drive.google.com/file/d/1H4bod9ZcS7XjV5tV5kZVfZiwk6Cj4YfD/view?usp=sharing

# STEP 4: now run the script using the command:
# python3 apeer.py

####################
### Main Program ###
####################
popdata = GetPopData()
chemdata = LoadPollutionData()

# now get county data
ofile = "2018_Toxics_Ambient_Concentrations.tract.tsv"
chemdata.to_csv(ofile, sep="\t")
cfile = GetCountyData(popdata, ofile)
chemdata_county = pd.read_csv(cfile, header=0, index_col=0, sep="\t")

# Individual chemicals can be selected and clustered:
#tract_list = ["METHANOL", "BENZOAPYRENE", "CRESOL_CRESYLIC_ACID_MIXED_ISOMERS", "ETHYLENE_GLYCOL", "4_AMINOBIPHENYL"]
#county_list = ["METHANOL", "ACETALDEHYDE", "ACROLEIN", "ACETONITRILE", "DIESEL_PM"]
#chemtract = chemdata[tract_list]
#chemcounty = chemdata_county[county_list]

chemtract = chemdata
chemcounty = chemdata_county

# get map data
county_json = LoadMapJSON("county")
tract_json = LoadMapJSON("tract")

# show maps for different clustering
lowclust = 2
hiclust = 5

for numclust in range(lowclust, hiclust):

	# Make a census-tract level map using the ShowClustersOnMap function - maps are saved in PNG format
	# parameter 1: mode (tract/cluster)
	# parameter 2: pollution data (pandas dataframe)
	# parameter 3: filename
	# parameter 4: json map data
	ShowClustersOnMap("tract", chemtract, numclust, "apeer_tract." + str(x), tract_json)

	# Make a county-level map using the ShowClustersOnMap function
	ShowClustersOnMap("county", chemcounty, numclust, "apeer_county." + str(x), county_json)


